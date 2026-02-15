"""
Excel export service using openpyxl with styled output.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(documents: list[dict], output_path: str) -> str:
    """
    Generate styled Excel from extracted data.
    documents: [{ filename, extracted_fields: [{key, label, value, confidence}] }]
    """
    wb = Workbook()

    # ---- Sheet 1: Extracted Data ----
    ws = wb.active
    ws.title = "Extracted Data"
    ws.sheet_properties.tabColor = "6C63FF"

    # Collect all unique field keys/labels across docs
    all_keys = []
    key_labels = {}
    for doc in documents:
        for f in doc.get("extracted_fields", []):
            if f["key"] not in key_labels:
                all_keys.append(f["key"])
                key_labels[f["key"]] = f["label"]

    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill(start_color="F0EEFF", end_color="F0EEFF", fill_type="solid")

    # Headers
    headers = ["Document"] + [key_labels.get(k, k) for k in all_keys]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, doc in enumerate(documents, 2):
        ws.cell(row=row_idx, column=1, value=doc["filename"]).border = thin_border

        field_map = {f["key"]: f for f in doc.get("extracted_fields", [])}
        for col_idx, key in enumerate(all_keys, 2):
            field = field_map.get(key, {})
            value = field.get("value", "—") or "—"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Alternate row color
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ---- Sheet 2: Confidence Scores ----
    ws2 = wb.create_sheet("Confidence Scores")
    ws2.sheet_properties.tabColor = "00C9A7"

    conf_header_fill = PatternFill(start_color="00C9A7", end_color="00C9A7", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = conf_header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws2.row_dimensions[1].height = 30

    for row_idx, doc in enumerate(documents, 2):
        ws2.cell(row=row_idx, column=1, value=doc["filename"]).border = thin_border
        field_map = {f["key"]: f for f in doc.get("extracted_fields", [])}
        for col_idx, key in enumerate(all_keys, 2):
            field = field_map.get(key, {})
            conf = field.get("confidence", 0)
            cell = ws2.cell(row=row_idx, column=col_idx, value=f"{conf}%" if conf else "—")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"📊 Excel generated: {output_path}")
    return output_path
